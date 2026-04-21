from openpyxl import load_workbook
import json

# Mở file Excel
wb = load_workbook("students_buoi6_02.xlsx")
ws = wb.active

# Header ở dòng 2
header_row = ws[2] 
headers = []
for cell in header_row:
    if cell.value is not None:
        headers.append(cell.value.strip())
    else:
        headers.append("") 

# Dữ liệu từ dòng 3
data = []
for row in ws.iter_rows(min_row=3, values_only=True):
    record = dict(zip(headers, row))
    # Lọc sinh viên điểm >= 9
    if record.get("Điểm", 0) >= 9:
        data.append(record)

# Ghi JSON
with open("report_buoi6_02.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

print("Đã tạo report_buoi6_02.json!")