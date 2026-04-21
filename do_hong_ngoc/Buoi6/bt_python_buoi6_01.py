#Bài tập 1

from openpyxl import load_workbook
import json

# Mở file Excel
wb = load_workbook("students_buoi6_01.xlsx")
ws = wb.active

# Lấy header (dòng đầu)
headers = []
for cell in ws[1]:
    headers.append(cell.value)

# Đọc dữ liệu
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    record = {}
    for i in range(len(headers)):
        record[headers[i]] = row[i]
    data.append(record)

# Ghi ra file JSON (giữ tiếng Việt)
with open("report_buoi6_01.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

print("Đã xuất report_buoi6_01.json!")